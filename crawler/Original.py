#先創建Excel，再將爬取資料進行處裡並一一填入
import requests
import bs4
from time import sleep
import openpyxl
import os

# 創建 Excel 檔案
wb = openpyxl.Workbook()
ws = wb.active

# 設定 Excel 標題
ws['A1'] = '職缺名稱'
ws['B1'] = '職缺連結'
ws['C1'] = '公司名稱'
ws['D1'] = '工作地區'
ws['E1'] = '薪資待遇'
ws['F1'] = '給薪方式'
ws['G1'] = '薪資下界'
ws['H1'] = '薪資上界'
ws['I1'] = '平均薪資'
ws['J1'] = '縣市'
ws['K1'] = '鄉鎮市區'

# 爬取的 URL
url = "https://www.104.com.tw/jobs/search/?isJobList=1&jobsource=joblist_search&order=15page=1&jobcat=2001001001&keyword=%E9%A3%9F%E5%93%81&area=6001001000,6001002000"
headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; Win64; x64; rv:47.0) Gecko/20210501 Firefox/47.0'}
try:
  res = requests.get(url,headers=headers)
  res.raise_for_status()
except Exception as e:
    print('錯誤訊息：', e)
else:
    res.encoding = 'utf-8'
    print(res.text)
    
soup = bs4.BeautifulSoup(res.text, "html.parser")  # 明確指定解析器
allJobsInform = soup.find_all('div', class_='info-container')

page = 1
while allJobsInform != []:
    print(f"=========================== 現在抓到第 {page} 頁資料 ===========================")
    for job in allJobsInform:
        try:
            # 抓取職缺資訊
            job_name = job.a.text.strip()
            job_link = "https:" + job.a['href']
            job_company = job.select('a')[1].text.strip()
            job_area = job.select('a')[3].string.strip()
            job_salary = job.select('a')[6].string.strip()

            job_county = job_area[:3]
            job_section = job_area[3:]
            PayWay = job_salary[:2]
            if PayWay == "待遇":
              PayWay = "面議"

            salary = ''
            for char in job_salary:
              if char.isdigit() or char=='~':
                salary += char

            if '~' in salary:
              lowEndSalary  = salary[:salary.find('~')]
              highEndSalary = salary[salary.find('~')+1:]
            else:
              lowEndSalary  = salary
              highEndSalary = salary

            if lowEndSalary != '' and highEndSalary != '':
              lowEndSalary = int(lowEndSalary)
              highEndSalary = int(highEndSalary)
              avgSalary = (lowEndSalary + highEndSalary)/2
            else:
              avgSalary = ''

            # 寫入 Excel
            print(job_name, job_link, job_company, job_area, job_salary, PayWay, lowEndSalary, highEndSalary, avgSalary, job_county, job_section)
            ws.append([job_name, job_link, job_company, job_area, job_salary, PayWay, lowEndSalary, highEndSalary, avgSalary, job_county, job_section])
        except Exception as e:
            print(f"資料解析錯誤：{e}")
    sleep(2)

    # 下一頁
    page += 1
    url = f"https://www.104.com.tw/jobs/search/?isJobList=1&jobsource=joblist_search&order=15&jobcat=2001001001&keyword=%E9%A3%90%E5%93%81&area=6001001000,6001002000&page={page}"
    res = requests.get(url, headers=headers) # Added headers here
    soup = bs4.BeautifulSoup(res.text, "html.parser")
    allJobsInform = soup.find_all('div', class_='info-container')

wb.save('./crawler/data/104CrawlResult.xlsx')
print(f"資料已儲存至: ./crawler/data/104CrawlResult.xlsx")
