#先創建pandas DataFrame，再將爬取資料進行處理並存入，最後一次性輸出到Excel
import requests
import bs4
from time import sleep
import pandas as pd
import xlwings as xw
import urllib.parse
swb = xw.Book.caller()

@xw.func
def scrape_jobs(keyword="", area="", jobcat="", max_page=0):
    try:
        # 創建 DataFrame
        columns = ['職缺名稱', '職缺連結', '公司名稱', '工作地區', '薪資待遇', 
                   '給薪方式', '薪資下界', '薪資上界', '平均薪資', '縣市', '鄉鎮市區']
        df = pd.DataFrame(columns=columns)

        # URL encode the keyword
        encoded_keyword = urllib.parse.quote(keyword)

        # 爬取的 URL
        url = f"https://www.104.com.tw/jobs/search/?jobsource=joblist_search&mode=s&order=15&page=1"
        if jobcat:
            url += f"&jobcat={jobcat}"
        if encoded_keyword:
            url += f"&keyword={encoded_keyword}"
        if area:
            url += f"&area={area}"

        res = requests.get(url)
        soup = bs4.BeautifulSoup(res.text, "html.parser")  # 明確指定解析器
        allJobsInform = soup.find_all('div', class_='info-container')

        page = 1
        while allJobsInform != [] and (max_page == 0 or page <= max_page):
            print(f"=========================== 現在抓到第 {page} 頁資料 ===========================")
            for job in allJobsInform:
                try:
                    # 抓取職缺資訊
                    job_name = job.a.text.strip()
                    job_link = job.a['href']
                    job_company = job.select('a')[1].text.strip()
                    job_area = job.select('a')[3].string.strip()
                    job_salary = job.select('a')[6].string.strip()

                    job_county = job_area[:3]
                    job_section = job_area[3:]
                    PayWay = job_salary[:2]
                    if PayWay == "待遇":
                      PayWay = "面議"

                    salary = ''
                    for char in job_salary:
                      if char.isdigit() or char=='~':
                        salary += char

                    if '~' in salary:
                      lowEndSalary  = salary[:salary.find('~')]
                      highEndSalary = salary[salary.find('~')+1:]
                    else:
                      lowEndSalary  = salary
                      highEndSalary = salary

                    if lowEndSalary != '' and highEndSalary != '':
                      lowEndSalary = int(lowEndSalary)
                      highEndSalary = int(highEndSalary)
                      avgSalary = (lowEndSalary + highEndSalary)/2
                    else:
                      avgSalary = ''

                    # 將資料加入 DataFrame
                    print(job_name, job_link, job_company, job_area, job_salary, PayWay, lowEndSalary, highEndSalary, avgSalary, job_county, job_section)
                    df.loc[len(df)] = [job_name, job_link, job_company, job_area, job_salary, PayWay, lowEndSalary, highEndSalary, avgSalary, job_county, job_section]
                except Exception as e:
                    print(f"資料解析錯誤：{e}")
            sleep(2)

            # 下一頁
            page += 1
            url = f"https://www.104.com.tw/jobs/search/?jobsource=joblist_search&mode=s&order=15&page={page}"
            if jobcat:
                url += f"&jobcat={jobcat}"
            if encoded_keyword:
                url += f"&keyword={encoded_keyword}"
            if area:
                url += f"&area={area}"

            res = requests.get(url)
            soup = bs4.BeautifulSoup(res.text, "html.parser")
            allJobsInform = soup.find_all('div', class_='info-container')

        # 寫入指定工作表
        swb.sheets['TEMP'].range('A1').value = url
        swb.sheets['職缺'].cells.clear_contents()
        swb.sheets['職缺'].range('A1').value = df.columns.tolist()
        swb.sheets['職缺'].range('A2').value =df.values

        return "爬蟲完成，資料已儲存至 ./data/104CrawlResult.xlsx"

    except Exception as e:
        swb.sheets['TEMP'].range('A2').value = f"Error: {e}"
        return f"爬蟲過程中發生錯誤: {e}"

def scrape_resume():
  #先創建Excel，再將爬取資料進行處裡並一一填入
  import requests
  import bs4
  from time import sleep
  import openpyxl
  import os

  # 創建 Excel 檔案
  wb = openpyxl.Workbook()
  ws = wb.active

  # 設定 Excel 標題
  ws['A1'] = '職缺名稱'
  ws['B1'] = '職缺連結'
  ws['C1'] = '公司名稱'
  ws['D1'] = '工作地區'
  ws['E1'] = '薪資待遇'
  ws['F1'] = '給薪方式'
  ws['G1'] = '薪資下界'
  ws['H1'] = '薪資上界'
  ws['I1'] = '平均薪資'
  ws['J1'] = '縣市'
  ws['K1'] = '鄉鎮市區'

  # 爬取的 URL
  url = "https://vip.104.com.tw/search/searchResult?loadTime=2025-04-28%2012%3A03%3A29&ec=1&kws=%E8%81%B7%E6%A5%AD%E5%AE%89%E5%85%A8%E8%A1%9B%E7%94%9F%E7%AE%A1%E7%90%86%E5%93%A1&jobcat=2009004008,2009004001&city=6001001000,6001002000&home=6001005000,6001006000&plastActionDateType=8&workExpTimeType=down&workExpTimeMin=3&workExpTimeMax=4&sex=2&empStatus=0&photo=1&updateDateType=1&contactPrivacy=0&sortType=RANK"
  res = requests.get(url)
  soup = bs4.BeautifulSoup(res.text, "html.parser")  # 明確指定解析器
  allJobsInform = soup.find_all('div', class_='vip-resume-card mb-2 resume-card')

  page = 1
  while allJobsInform != []:
      print(f"=========================== 現在抓到第 {page} 頁資料 ===========================")
      # for job in allJobsInform:
      #     try:
      #         # 抓取職缺資訊
      #         job_name = job.a.text.strip()
      #         job_link = "https:" + job.a['href']
      #         job_company = job.select('a')[1].text.strip()
      #         job_area = job.select('a')[3].string.strip()
      #         job_salary = job.select('a')[6].string.strip()

      #         job_county = job_area[:3]
      #         job_section = job_area[3:]
      #         PayWay = job_salary[:2]
      #         if PayWay == "待遇":
      #           PayWay = "面議"

      #         salary = ''
      #         for char in job_salary:
      #           if char.isdigit() or char=='~':
      #             salary += char

      #         if '~' in salary:
      #           lowEndSalary  = salary[:salary.find('~')]
      #           highEndSalary = salary[salary.find('~')+1:]
      #         else:
      #           lowEndSalary  = salary
      #           highEndSalary = salary

      #         if lowEndSalary != '' and highEndSalary != '':
      #           lowEndSalary = int(lowEndSalary)
      #           highEndSalary = int(highEndSalary)
      #           avgSalary = (lowEndSalary + highEndSalary)/2
      #         else:
      #           avgSalary = ''

      #         # 寫入 Excel
      #         print(job_name, job_link, job_company, job_area, job_salary, PayWay, lowEndSalary, highEndSalary, avgSalary, job_county, job_section)
      #         ws.append([job_name, job_link, job_company, job_area, job_salary, PayWay, lowEndSalary, highEndSalary, avgSalary, job_county, job_section])
      #     except Exception as e:
      #         print(f"資料解析錯誤：{e}")
      sleep(2)

      # 下一頁
      page += 1
      url = f"https://www.104.com.tw/jobs/search/?jobsource=index_s&keyword=%E5%A4%A7%E6%95%B8%E6%93%9A&mode=s&page={page}"
      res = requests.get(url)
      soup = bs4.BeautifulSoup(res.text, "html.parser")
      allJobsInform = soup.find_all('div', class_='info-container')
      # wb.save('./data/104CrawlResult.xlsx')

if __name__ == "__main__":
    # This block allows you to run the script directly for testing
    scrape_resume()
